import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json


class ExcelData:
    def __init__(self, file_name: str):
        self.file_name = file_name
        self.full_file_path = f'excel/{file_name}.xlsx'

    def __create_excel_file(self):
        """Creating file with data"""
        wb = openpyxl.Workbook()
        wb.save(self.full_file_path)
        workbook = load_workbook(filename=self.full_file_path)
        workbook.worksheets[0].title = 'data'
        wb.save(self.full_file_path)
        return workbook

    def __get_json_data(self):
        """Get data from json"""
        try:
            with open(f'json/{self.file_name}.json', 'r', encoding='utf-8') as file:
                return json.loads(file.read())
        except Exception as ex:
            print(f'[FATAL] Cant open json file with data. Script shutdown,\n{ex}')
            sys.exit()

    def write_data_to_excel(self):
        """Write data to excel"""
        workbook = self.__create_excel_file()
        src_file = self.__get_json_data()

        sheet = 'data'
        excel_column = 1
        excel_row = 1

        # write headers
        color_code = '9FC5E8'
        headers_list = src_file[0]
        for header in headers_list:
            workbook[sheet].cell(row=excel_row, column=excel_column).value = header
            workbook[sheet].cell(row=excel_row, column=excel_column).fill = PatternFill(fgColor=color_code,
                                                                                        fill_type='solid')
            excel_column += 1

        # increase column width
        columns_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        for column in columns_letter:
            workbook[sheet].column_dimensions[column].width = 20

        # write values
        excel_row = 2
        for data in src_file:
            for num, each in enumerate(data):
                workbook[sheet].cell(row=excel_row, column=num + 1).value = data[each]
                if num >= 6:  # max users columns
                    break
            excel_row += 1

        workbook.save(self.full_file_path)
        workbook.close()
